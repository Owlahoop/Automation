# Import Package
import openpyxl

wk = openpyxl.load_workbook("C:\\Users\\Ryan\\Desktop\\TestData.xlsx")
sh = wk["Sheet1"]

# Find rows with data
rows = sh.max_row
columns = sh.max_column

print("Total rows are " + str(rows))
print("Total columns are " + str(columns))

for r in sh['A1':'C4']:
    for c in r:
        print(c.value)

# for i in range(1, rows + 1):
#    for j in range(1, columns + 1):
#        c = sh.cell(i, j)
#        print(c.value)