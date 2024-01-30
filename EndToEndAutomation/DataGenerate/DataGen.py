import openpyxl


def data_generator():
    wk = openpyxl.load_workbook("C:\\Users\\Ryan\\Desktop\\TData10.xlsx")
    sh = wk['Sheet1']
    r = sh.max_row
    li = []
    # Excel loop
    for i in range(1, r+1):
        # Empty list
        li1 = []
        # Assigns cell info to variables un and up (row,column)
        un = sh.cell(i, 1)
        up = sh.cell(i, 2)
        # Inserts the variable values into the empty list
        li1.insert(0, un.value)
        li1.insert(1, up.value)
        # Transfers list to another variable, so it is not overwritten as loop continues
        li.insert(i-1, li1)
    return li
